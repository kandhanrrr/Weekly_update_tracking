# pyright: reportMissingTypeStubs=false
"""
Weekly Update Tracker — Automated Reminder & Follow-up System
==============================================================
Reads a configurable sheet from any Excel tracker file and:
  - Sends a reminder email 2 days before ETA for WIP / Not yet started tasks
  - Sends daily follow-up emails when ETA is exceeded and task is still open
  - Silently skips tasks whose status has been updated to any closed state
    (Done, Dropped, NA, Review Done/Closed) even if ETA was already exceeded
  - Generates a Friday weekly summary email to the team

All settings (Excel file, sheet name, SMTP, owners) are in config.ini.
No code changes are needed for normal team setup.

Usage:
    python weekly_update_tracker.py
            # dry-run (print actions, no emails)
    python weekly_update_tracker.py --send
            # send real emails via SMTP
    python weekly_update_tracker.py --weekly-summary
            # force weekly summary regardless of day
"""

import argparse
import configparser
import re
import smtplib
import sys
import warnings
from collections import defaultdict
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]

# ---------------------------------------------------------------------------
# Load configuration from config.ini (same folder as this script)
# Password is read from Windows Credential Manager — never stored in a file
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent
CONFIG_FILE = BASE_DIR / "config.ini"

# Read tracker-specific settings from config.ini [TRACKER] section
_raw_cfg = configparser.ConfigParser()
_raw_cfg.read(CONFIG_FILE)
_excel_rel = _raw_cfg.get(
    "TRACKER",
    "excel_file",
    fallback="PTH_progress_tracking.xlsx",
)
_sheet = _raw_cfg.get("TRACKER", "sheet_name", fallback="auto")
CRED_SERVICE = _raw_cfg.get(
    "TRACKER",
    "credential_service",
    fallback="Weekly_Update_Tracker",
)

EXCEL_FILE = (BASE_DIR / _excel_rel).resolve()
SHEET_NAME = _sheet


def _load_password_from_credential_manager(username: str) -> str:
    """
    Retrieve the email password from Windows Credential Manager.
    Returns empty string if not found (will fall back to config.ini smtp_pass).
    """
    try:
        import win32cred  # type: ignore[import-untyped]

        cred = win32cred.CredRead(CRED_SERVICE, win32cred.CRED_TYPE_GENERIC)
        return cred["CredentialBlob"]
    except ImportError:
        print(
            "[WARN] win32cred not available - "
            "falling back to config.ini smtp_pass"
        )
        return ""
    except Exception:
        return ""


def _load_config():
    """Read config.ini and return (smtp_cfg, owner_map, team_email)."""
    if not CONFIG_FILE.exists():
        print(f"[ERROR] config.ini not found at {CONFIG_FILE}")
        print("        Please create it using the template provided.")
        sys.exit(1)

    cfg = configparser.ConfigParser()
    cfg.read(CONFIG_FILE)

    smtp_user = cfg.get("EMAIL", "smtp_user", fallback="")

    # Password priority: Windows Credential Manager > config.ini smtp_pass
    password = _load_password_from_credential_manager(smtp_user)
    if not password:
        password = cfg.get("EMAIL", "smtp_pass", fallback="")
        if password and password != "YOUR_PASSWORD_HERE":
            print(
                "[INFO] Using password from config.ini "
                "(consider running save_password.py for better security)"
            )
        else:
            password = ""

    smtp = {
        "host": cfg.get("EMAIL", "smtp_host", fallback="smtp.intel.com"),
        "port": cfg.getint("EMAIL", "smtp_port", fallback=587),
        "user": smtp_user,
        "pass": password,
    }
    team_email = cfg.get("EMAIL", "team_email", fallback="")

    return smtp, team_email


_SMTP_CFG, TEAM_EMAIL = _load_config()

SMTP_HOST = _SMTP_CFG["host"]
SMTP_PORT = _SMTP_CFG["port"]
SMTP_USER = _SMTP_CFG["user"]
SMTP_PASS = _SMTP_CFG["pass"]


# Build owner map — configparser lowercases keys, so re-read preserving case
def _build_owner_map(team_email: str) -> dict:
    """Re-read config.ini preserving original casing for owner names."""

    class _CaseConfigParser(configparser.RawConfigParser):
        def optionxform(self, optionstr: str) -> str:
            return optionstr

    cfg = _CaseConfigParser()
    cfg.read(CONFIG_FILE)
    owner_map = {}
    if cfg.has_section("OWNERS"):
        for name, email in cfg.items("OWNERS"):
            owner_map[name.strip()] = email.strip() if email.strip() else None
    # Fallback for generic entries
    owner_map.setdefault("All", team_email)
    owner_map.setdefault("Team", team_email)
    owner_map.setdefault("TBD", None)
    return owner_map


OWNER_EMAIL_MAP = _build_owner_map(TEAM_EMAIL)

# ---------------------------------------------------------------------------
# Status constants
# ---------------------------------------------------------------------------

ACTIVE_STATUSES = {"wip", "not yet started", "open"}
CLOSED_STATUSES = {"done", "dropped", "na", "review done/closed"}

# Auto-normalise common informal/variant status labels to canonical values.
# Teams often use mixed phrasing; this avoids rejecting valid rows.
STATUS_ALIASES: dict[str, str] = {
    # → WIP
    "in progress":       "wip",
    "in-progress":       "wip",
    "inprogress":        "wip",
    "active":            "wip",
    "ongoing":           "wip",
    "started":           "wip",
    "in review":         "wip",
    # → Not yet started
    "not started":       "not yet started",
    "not_yet_started":   "not yet started",
    "todo":              "not yet started",
    "to do":             "not yet started",
    "to-do":             "not yet started",
    "backlog":           "not yet started",
    "planned":           "not yet started",
    # → Open
    "pending":           "open",
    "new":               "open",
    "reopened":          "open",
    "re-opened":         "open",
    "re opened":         "open",
    # → Done
    "completed":         "done",
    "complete":          "done",
    "finished":          "done",
    "done - verified":   "done",
    "done/closed":       "done",
    "closed":            "done",
    "resolved":          "done",
    # → Dropped
    "cancel":            "dropped",
    "cancelled":         "dropped",
    "canceled":          "dropped",
    "not required":      "dropped",
    "deferred":          "dropped",
    # → NA
    "n/a":               "na",
    "not applicable":    "na",
    "no action":         "na",
    "not needed":        "na",
    # → Review Done/Closed
    "review done":       "review done/closed",
    "review complete":   "review done/closed",
    "review closed":     "review done/closed",
}

SUMMARY_STATUS_ORDER = [
    "completed",
    "open",
    "wip",
    "not yet started",
    "dropped",
    "na",
    "review done/closed",
    "other",
]

HEADER_ALIASES = {
    "module": {"module", "area", "functional area", "team"},
    "task": {"task", "item", "title", "description", "work item"},
    "owner": {"owner", "assignee", "responsible", "person", "lead"},
    "priority": {"priority", "prio"},
    "status": {"status", "state", "progress"},
    "eta": {"eta", "due", "due date", "target", "target date", "deadline"},
    "remark": {
        "remark",
        "remarks",
        "notes",
        "note",
        "comment",
        "comments",
        "update",
    },
}

# ---------------------------------------------------------------------------
# Work-week helpers
# ---------------------------------------------------------------------------


def ww_to_date(ww_str: str, year: int | None = None) -> datetime | None:
    """
    Convert an Intel/ISO work-week string to the Friday (end of week) datetime.
    Recognised formats: WW30, ww30, WW30p5, WW25p3, WW24P1
    Returns None if the string cannot be parsed.
    """
    if not ww_str:
        return None
    s = str(ww_str).strip().upper()
    match = re.search(r"WW(\d{1,2})(?:P(\d))?", s)
    if not match:
        return None

    ww_num = int(match.group(1))
    part = int(match.group(2)) if match.group(2) else 5  # default to Friday

    if year is None:
        year = datetime.today().year

    # ISO week-1 Monday = Monday of the week that contains Jan 4
    jan4 = datetime(year, 1, 4)
    week1_monday = jan4 - timedelta(days=jan4.weekday())
    target_monday = week1_monday + timedelta(weeks=ww_num - 1)
    # part: 1=Mon … 5=Fri … 7=Sun
    day_offset = min(part - 1, 4)  # cap at Friday
    return target_monday + timedelta(days=day_offset)


def current_ww() -> tuple[int, int]:
    """Return (year, iso_week) for today."""
    today = datetime.today()
    iso = today.isocalendar()
    return iso[0], iso[1]


def _normalize_header(value) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _build_header_map(sheet) -> tuple[dict[str, int], int]:
    """
    Find the best header row in a sheet and map expected logical fields
    to column indices.

    Returns (header_map, header_row_index). The map contains only
    the keys that were found.
    """
    best_map: dict[str, int] = {}
    best_row = 1
    best_score = 0

    rows = sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row, 15),
        values_only=True,
    )
    for row_idx, row in enumerate(rows, start=1):
        header_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            header = _normalize_header(cell)
            for logical_name, aliases in HEADER_ALIASES.items():
                if header in aliases:
                    header_map[logical_name] = col_idx

        score = len(set(header_map) & {"task", "status", "eta"})
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_map = header_map

    if best_score < 2:
        raise ValueError(
            "Could not detect a usable header row. "
            "Expected at least Task/Status/ETA columns."
        )

    return best_map, best_row


def _select_sheet(workbook):
    """
    Return the worksheet to process.

    If a sheet name is configured and exists, use it.
    Otherwise search all sheets for one with a recognizable header row.
    """
    configured = str(SHEET_NAME).strip()
    if configured and configured.lower() not in {"auto", "*", "any"}:
        if configured in workbook.sheetnames:
            return workbook[configured]
        print(
            f"[WARN] Sheet '{configured}' not found — "
            "auto-detecting from workbook tabs."
        )

    for ws in workbook.worksheets:
        try:
            _build_header_map(ws)
            return ws
        except ValueError:
            continue

    raise ValueError(
        "No sheet with recognizable Task/Status/ETA headers was found."
    )


# ---------------------------------------------------------------------------
# Email helpers
# ---------------------------------------------------------------------------

def _resolve_owner_emails(owner_field: str) -> list[str]:
    """
    owner_field may be 'Shivagiri', 'Kandhan/Shivagiri', 'Partap/Anju', etc.
    Returns a list of email addresses, skipping None entries.
    """
    if not owner_field:
        return []
    parts = re.split(r"[/,&]", str(owner_field).strip())
    emails = []
    for part in parts:
        name = part.strip()
        email = OWNER_EMAIL_MAP.get(name)
        if email:
            emails.append(email)
        elif name not in ("TBD", "All", "Team"):
            # Unknown owner — flag it
            print(f"  [WARN] No email mapped for owner '{name}'")
    return emails


def _send_email(to: list[str], subject: str, body: str, dry_run: bool = True):
    if not to:
        print("  [SKIP] No recipients resolved — skipping email.")
        return
    print(f"\n{'[DRY-RUN] ' if dry_run else ''}EMAIL")
    print(f"  To     : {', '.join(to)}")
    print(f"  Subject: {subject}")
    print(f"  Body   :\n{body}")
    if dry_run:
        return

    msg = MIMEMultipart("alternative")
    msg["From"] = SMTP_USER
    msg["To"] = ", ".join(to)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            if SMTP_PORT != 25:
                server.starttls()
            if SMTP_PASS and SMTP_PORT != 25:
                server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, to, msg.as_string())
        print("  [SENT]")
    except Exception as exc:
        print(f"  [ERROR] Failed to send: {exc}")


# ---------------------------------------------------------------------------
# Email templates
# ---------------------------------------------------------------------------

def _remind_email(task: dict, days_left: int) -> tuple[str, str]:
    subject = (
        "[Weekly Update Reminder] "
        f"ETA in {days_left} day(s): {task['task'][:60]}"
    )
    body = (
        f"Hi {task['owner']},\n\n"
        "This is a friendly reminder that the following task "
        f"is due in {days_left} day(s).\n\n"
        f"  Task    : {task['task']}\n"
        f"  Status  : {task['status']}\n"
        f"  ETA     : {task['eta']}\n"
        f"  Remarks : {task['remark'] or 'N/A'}\n\n"
        f"Please update the tracker or reach out if you need support.\n\n"
        f"Regards,\nWeekly Update Tracker"
    )
    return subject, body


def _overdue_email(task: dict, days_overdue: int) -> tuple[str, str]:
    subject = f"[Weekly Update Overdue] Action needed: {task['task'][:60]}"
    body = (
        f"Hi {task['owner']},\n\n"
        "The following task has exceeded its ETA by "
        f"{days_overdue} day(s) and is still open.\n\n"
        f"  Task     : {task['task']}\n"
        f"  Status   : {task['status']}\n"
        f"  ETA      : {task['eta']}\n"
        f"  Remarks  : {task['remark'] or 'N/A'}\n\n"
        f"Please update the status in the tracker immediately.\n"
        "If this task is complete/dropped, "
        "change the status so reminders stop.\n\n"
        f"Regards,\nWeekly Update Tracker"
    )
    return subject, body


def _weekly_summary_email(tasks: list[dict]) -> tuple[str, str]:
    today_str = datetime.today().strftime("%Y-%m-%d")
    subject = f"[Weekly Update Summary] Open Tasks as of {today_str}"

    wip_tasks = [t for t in tasks if t["status"].lower() == "wip"]
    pending_tasks = [
        t for t in tasks if t["status"].lower() == "not yet started"
    ]

    def fmt_tasks(lst):
        if not lst:
            return "  (none)\n"
        lines = []
        for t in lst:
            lines.append(
                f"  • [{t['owner']}] {t['task'][:70]}"
                f"  |  ETA: {t['eta'] or 'TBD'}"
            )
        return "\n".join(lines) + "\n"

    body = (
        f"Hi Team,\n\n"
        "Here is the weekly status of open items in "
        f"the tracker ({today_str}).\n\n"
        f"--- WIP ({len(wip_tasks)} tasks) ---\n"
        f"{fmt_tasks(wip_tasks)}\n"
        f"--- Not Yet Started ({len(pending_tasks)} tasks) ---\n"
        f"{fmt_tasks(pending_tasks)}\n"
        f"Please review and update your tasks in the tracker.\n\n"
        f"Regards,\nWeekly Update Tracker"
    )
    return subject, body


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------

def load_active_tasks(sheet) -> list[dict]:
    """
    Load only WIP and Not yet started tasks from the sheet.
    Tasks with closed statuses are excluded entirely — no emails sent for them.
    """
    tasks = []
    header_map, header_row = _build_header_map(sheet)
    required_fields = {"task", "status", "eta"}
    missing = sorted(required_fields - set(header_map))
    if missing:
        raise ValueError(
            f"Missing required column(s): {', '.join(missing)}. "
            f"Found headers: {sorted(header_map)}"
        )

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        status_idx = header_map["status"]
        has_status = status_idx < len(row) and row[status_idx]
        status_raw = str(row[status_idx]).strip() if has_status else ""
        status_norm = STATUS_ALIASES.get(
            status_raw.lower(), status_raw.lower()
        )

        if status_norm not in ACTIVE_STATUSES:
            continue  # closed / dropped / NA — skip silently

        task_idx = header_map["task"]
        owner_idx = header_map.get("owner")
        eta_idx = header_map["eta"]
        remark_idx = header_map.get("remark")
        prio_idx = header_map.get("priority")
        module_idx = header_map.get("module")

        def _cell(idx: int | None, default=""):
            if idx is None or idx >= len(row):
                return default
            value = row[idx]
            return value if value is not None else default

        owner_val = _cell(owner_idx, "TBD")
        tasks.append({
            "module": _cell(module_idx, ""),
            "task": str(_cell(task_idx, "(no task name)")).strip(),
            "owner": str(owner_val).strip() if owner_val else "TBD",
            "status": status_raw,
            "eta": str(_cell(eta_idx, "")).strip(),
            "remark": str(_cell(remark_idx, "")).strip(),
            "priority": str(_cell(prio_idx, "")).strip(),
        })
    return tasks


def _status_for_summary(status_raw: str) -> str:
    """Normalize status into summary buckets."""
    norm = STATUS_ALIASES.get(status_raw.lower(), status_raw.lower())
    if norm == "done":
        return "completed"
    if norm in {
        "open",
        "wip",
        "not yet started",
        "dropped",
        "na",
        "review done/closed",
    }:
        return norm
    return "other"


def build_module_status_summary(sheet) -> list[dict[str, int | str]]:
    """Build module-wise status counts across all task rows in the sheet."""
    header_map, header_row = _build_header_map(sheet)
    task_idx = header_map.get("task")
    status_idx = header_map.get("status")
    module_idx = header_map.get("module")

    if task_idx is None or status_idx is None:
        return []

    by_module: dict[str, dict[str, int]] = defaultdict(
        lambda: {k: 0 for k in SUMMARY_STATUS_ORDER}
    )

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if task_idx >= len(row):
            continue
        task_val = row[task_idx]
        if task_val is None or str(task_val).strip() == "":
            continue

        status_raw = ""
        if status_idx < len(row) and row[status_idx] is not None:
            status_raw = str(row[status_idx]).strip()
        status_bucket = _status_for_summary(status_raw)

        module = "Unassigned"
        if module_idx is not None and module_idx < len(row) and row[module_idx]:
            module = str(row[module_idx]).strip() or "Unassigned"

        by_module[module][status_bucket] += 1

    out: list[dict[str, int | str]] = []
    for module in sorted(by_module):
        counts = by_module[module]
        total = sum(counts.values())
        out.append(
            {
                "module": module,
                "completed": counts["completed"],
                "open": counts["open"],
                "wip": counts["wip"],
                "not_yet_started": counts["not yet started"],
                "dropped": counts["dropped"],
                "total": total,
            }
        )
    return out


def print_module_status_summary(sheet) -> None:
    """Print module-wise status table for weekly reporting output."""
    rows = build_module_status_summary(sheet)
    print("\n=== Module-wise Status Summary ===")
    print("| Module | Completed | Open | WIP | Not yet started | Dropped | Total |")
    print("|---|---:|---:|---:|---:|---:|---:|")
    if not rows:
        print("| (none) | 0 | 0 | 0 | 0 | 0 | 0 |")
        return
    for r in rows:
        print(
            "| {module} | {completed} | {open} | {wip} | {not_yet_started} | {dropped} | {total} |".format(
                module=r["module"],
                completed=r["completed"],
                open=r["open"],
                wip=r["wip"],
                not_yet_started=r["not_yet_started"],
                dropped=r["dropped"],
                total=r["total"],
            )
        )


def process_reminders(tasks: list[dict], dry_run: bool = True):
    """
    For each active task:
      - If ETA is within the next 2 calendar days → send reminder
      - If ETA has already passed              → send overdue follow-up
      - If ETA is TBD / unparseable            → log a warning
    """
    today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    reminders_sent = 0
    overdue_sent = 0
    tbd_count = 0

    for task in tasks:
        eta_date = ww_to_date(task["eta"])

        if eta_date is None:
            tbd_count += 1
            print(
                f"  [TBD ETA] '{task['task'][:55]}' "
                f"(owner: {task['owner']})"
            )
            continue

        eta_date = eta_date.replace(hour=0, minute=0, second=0, microsecond=0)
        delta = (eta_date - today).days  # positive = future, negative = past

        recipients = _resolve_owner_emails(task["owner"])

        if 0 <= delta <= 2:
            # ETA is today or within 2 days
            subject, body = _remind_email(task, delta)
            _send_email(recipients, subject, body, dry_run=dry_run)
            reminders_sent += 1

        elif delta < 0:
            # ETA already exceeded — task still open
            subject, body = _overdue_email(task, abs(delta))
            _send_email(recipients, subject, body, dry_run=dry_run)
            overdue_sent += 1

        else:
            print(
                f"  [OK] '{task['task'][:55]}' | "
                f"ETA: {task['eta']} ({delta} days away)"
            )

    print("\n--- Summary ---")
    print(f"  Active tasks processed  : {len(tasks)}")
    print(f"  Reminders queued (<=2d) : {reminders_sent}")
    print(f"  Overdue follow-ups      : {overdue_sent}")
    print(f"  TBD/unparseable ETAs    : {tbd_count}")


def process_weekly_summary(
    tasks: list[dict],
    dry_run: bool = True,
    force: bool = False,
):

    """Send weekly summary every Friday (or on --weekly-summary flag)."""
    today = datetime.today()
    is_friday = (today.weekday() == 4)  # 4 = Friday

    if not is_friday and not force:
        print(
            "\n[INFO] Today is not Friday - skipping weekly summary. "
            "Use --weekly-summary to force."
        )
        return

    subject, body = _weekly_summary_email(tasks)
    _send_email([TEAM_EMAIL], subject, body, dry_run=dry_run)


# ---------------------------------------------------------------------------
# AR upsert — write extracted AR rows into AR_Tracking sheet
# ---------------------------------------------------------------------------

AR_SHEET_NAME = "AR_Tracking"
AR_HEADERS = ["Date", "Channel", "Task", "Owner", "Status", "ETA", "Remark", "Source"]


def _get_or_create_ar_sheet(wb):
    """Return the AR_Tracking worksheet, creating it with headers if absent.
    If the sheet exists but is missing the Channel column, add it.
    """
    if AR_SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(AR_SHEET_NAME)
        ws.append(AR_HEADERS)
        return ws
    ws = wb[AR_SHEET_NAME]
    # Backfill Channel column if sheet was created before it existed
    header = [c.value for c in ws[1]]
    if "Channel" not in header:
        insert_col = 2  # after Date
        ws.insert_cols(insert_col)
        ws.cell(row=1, column=insert_col).value = "Channel"
    return ws


def _build_source_index(ws) -> dict[str, int]:
    """Return {source_key: row_number} for all existing rows in the AR sheet."""
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return {}
    try:
        src_col = list(header_row).index("Source")
    except ValueError:
        return {}
    index: dict[str, int] = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if src_col < len(row) and row[src_col]:
            index[str(row[src_col])] = row_num
    return index


def upsert_ar_rows(json_path: str | Path) -> None:
    """
    Read normalized AR rows from *json_path* and upsert into AR_Tracking sheet.
    - Dedup key: source_key field (SHA of source+date+task).
    - Existing key: update Status, ETA, Remark columns.
    - New key: append row.
    Prints a summary table on completion.
    """
    import json as _json

    rows: list[dict[str, Any]] = _json.loads(
        Path(json_path).read_text(encoding="utf-8")
    )

    warnings.filterwarnings(
        "ignore",
        message="Data Validation extension is not supported and will be removed",
        category=UserWarning,
    )
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = _get_or_create_ar_sheet(wb)

    # Column index map from current headers (1-based for openpyxl cell())
    header = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(header) if h}

    src_index = _build_source_index(ws)

    inserted = updated = skipped = 0

    for r in rows:
        key = r.get("source_key", "")
        task = r.get("task", "")
        owner = r.get("owner", "TBD")
        status = r.get("status", "Open")
        eta = r.get("eta", "TBD")
        channel = r.get("folder") or r.get("source", "")  # e.g. "Outlook:Inbox", "Outlook:Sent", "Teams"
        remark = r.get("summary", "")
        if r.get("sender"):
            remark = f"{remark} | From: {r['sender']}".strip(" |")
        date_str = r.get("date", "")

        if not task:
            skipped += 1
            continue

        if key and key in src_index:
            # Update existing row — also backfill Channel if it was empty
            row_num = src_index[key]
            for col_name, value in [
                ("Channel", channel),
                ("Status", status),
                ("ETA", eta),
                ("Remark", remark),
            ]:
                if col_name in col_idx:
                    ws.cell(row=row_num, column=col_idx[col_name]).value = value
            updated += 1
        else:
            # Append new row in header column order
            new_row = [
                date_str,
                channel,
                task,
                owner,
                status,
                eta,
                remark,
                key,
            ]
            ws.append(new_row)
            if key:
                src_index[key] = ws.max_row
            inserted += 1

    wb.save(EXCEL_FILE)

    print(f"\n# AR Upsert Summary — {Path(EXCEL_FILE).name} / {AR_SHEET_NAME}")
    print("| Metric | Count |")
    print("|---|---|")
    print(f"| Inserted (new) | {inserted} |")
    print(f"| Updated (existing) | {updated} |")
    print(f"| Skipped (no task) | {skipped} |")
    print(f"| Total processed | {len(rows)} |")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Weekly Update Task Reminder Automation"
    )
    parser.add_argument(
        "--send",
        action="store_true",
        help="Send real emails (default: dry-run)",
    )
    parser.add_argument(
        "--weekly-summary",
        action="store_true",
        help="Force weekly summary email regardless of day",
    )
    parser.add_argument(
        "--upsert-ar",
        metavar="JSON_PATH",
        default=None,
        help="Path to AR rows JSON (from parse_pending_from_graph_exports.py --output-json). "
             "Upserts rows into AR_Tracking sheet and exits.",
    )
    args = parser.parse_args()

    if args.upsert_ar:
        upsert_ar_rows(args.upsert_ar)
        return

    dry_run = not args.send

    now_str = datetime.today().strftime("%Y-%m-%d %H:%M")
    mode_str = (
        "LIVE (sending emails)"
        if not dry_run
        else "DRY-RUN (no emails sent)"
    )
    print(f"Weekly Update Tracker - {now_str}")
    print(f"Mode: {mode_str}")
    print(f"File: {EXCEL_FILE}\n")

    if not EXCEL_FILE.exists():
        print(f"[ERROR] Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    # Suppress openpyxl data-validation extension warning from some workbooks.
    warnings.filterwarnings(
        "ignore",
        message=(
            "Data Validation extension is not supported and will be removed"
        ),
        category=UserWarning,
    )
    wb = openpyxl.load_workbook(EXCEL_FILE)
    try:
        ws = _select_sheet(wb)
    except ValueError as exc:
        print(f"[ERROR] {exc}")
        print(f"        Available sheets: {wb.sheetnames}")
        sys.exit(1)

    tasks = load_active_tasks(ws)

    print(f"Loaded {len(tasks)} active task(s) from '{ws.title}'.\n")
    print("=== Processing ETA Reminders ===")
    process_reminders(tasks, dry_run=dry_run)
    print_module_status_summary(ws)

    print("\n=== Processing Weekly Summary ===")
    process_weekly_summary(
        tasks,
        dry_run=dry_run,
        force=args.send and args.weekly_summary,
    )


if __name__ == "__main__":
    main()
